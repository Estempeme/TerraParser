#!/usr/bin/env python
# coding: utf8

import unittest
import sys
from openpyxl import load_workbook
import datetime
# import pdftotext
from fpdf import FPDF
# import time
# import os
import pdb
# import pickle
from Artikel import Artikel
from mwst import Mwst


class Bestellung:
    def __init__(self, liste, datum):

        self.liste = liste
        if datum == None:
            self.datum = datetime.date(0, 0, 0)
        else:
            self.datum = datum
        self.kommentar = "…………………………………………………\nKommentar:\n"

    def __make_outputstring(self):
        """
        setzt den String für das Speichern in Datei self.speichern()
        oder Druckausgabe self.drucken() zusammen
        return outputstring
        """
        outputstring = f"Terrabestellung vom {self.datum.day}.{self.datum.month}.\n\n"
        outputstring += '{:^25}'.format('Artikelname')
        outputstring += '{:^12}'.format('Netto')
        outputstring += '{:^7}'.format('MwSt ' + "{:1.0f}".format(Mwst.erm*100) + '%')
        outputstring += '{:^7}'.format("{:2.0f}".format(Mwst.norm*100) + '%')
        outputstring += '\n…………………………………………………………………………………………………………………………………\n'

        käse = False

        for i in range(len(self.liste)):
            artikel = self.liste[i]

            käse = 190000 > int(artikel.artikelnummer) > 130844

            outputstring += '{:.<25.25}'.format(artikel.bezeichnung)
            outputstring += '{:>8.2f}'.format(artikel.grundpreis).replace('.', ',')

            if käse:
                outputstring += '{:>8.2f}'.format(Mwst.berechnen(artikel.grundpreis, Mwst.erm)).replace('.', ',')
                outputstring += '     /100 Gramm'
            else:
                outputstring += '{:>8.2f}'.format(Mwst.berechnen(artikel.grundpreis, Mwst.erm)).replace('.', ',')
                outputstring += '{:>8.2f}'.format(Mwst.berechnen(artikel.grundpreis, Mwst.norm)).replace('.', ',')

            outputstring += '\n'

        outputstring += self.kommentar
        return outputstring

    def auf_konsole_drucken(self):
        print(self.__make_outputstring())

    def als_text_speichern(self):
        outputfile_name = 'Preise_Terrabestellung_' + self.datum.isoformat() + '.txt'
        try:
            outputfile = open(outputfile_name, "x")
            outputfile.write("Hallo\n")
            outputfile.write(self.__make_outputstring())
            print(outputfile_name + ' wird angelegt.')
            outputfile.close()
        except:
            print("Datei schon vorhanden, kein Speichern erfolgt.")
            """
            TODO:
                Abfrage, ob Datei überschrieben werden soll oder unter
                anderem Namen gespeichert
            """


    def als_pdf_speichern(bestellung):
        """
        Die Funktion schreibe_pdf(bestellung) nimmt ein Bestellungsobjekt entgegen
        und erzeugt daraus eine pdf-Datei mit dem Namen 'Terrabestellung vom + datum.
        Diese pdf-Datei wird ohne Rückfrage im Arbeitsverzeichnis gespeichert.
        """
        bestell_liste = bestellung.liste
        datum = bestellung.datum.strftime("%A, %d. %b. %Y")
        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.add_page()

        # Hier wird das logo gesetzt
        try:
            pdf.image('logo-fcn.jpg', x=160, y=8, w=30)
        except:
            print("Logo nicht gefunden")

        # Schriftart und Größe der Überschrift
        pdf.set_font("Times", size=14)

        # Titel
        pdf.cell(180, 20, txt="Terrabestellung vom " + datum, ln=1, align="C")

        pdf.set_line_width(.5)

        pdf.line(10, 37, 190, 37)
        pdf.ln(20)  # move 20 down

        # Schriftart und Größe des Textes
        pdf.set_font("Times", size=12)

        # Überschrift erst Zeile
        pdf.cell(65, 8, txt="Bezeichnung", align="C")
        pdf.cell(20, 8, txt="Grundpreis", align="C")
        pdf.cell(20, 8, txt=f"{Mwst.erm*100:.0f} %", align="C")
        pdf.cell(20, 8, txt=f"{Mwst.norm*100:.0f} %", align="C")
        pdf.cell(20, 8, txt=f"{Mwst.erm*100:.0f} %/100g", ln=1, align="R")

        pdf.set_font("Times", size=10)

        for artikel in bestell_liste:
            # zeile = 0
            pro100gramm = Mwst.berechnen(artikel.grundpreis/10, Mwst.erm)
            pdf.cell(60, 8, txt=artikel.bezeichnung, align="L")
            pdf.cell(20, 8, txt=f"{artikel.grundpreis:.2f}".replace('.', ','), align="R")
            pdf.cell(20, 8, txt=f"{artikel.mwstErm:.2f}".replace('.', ','),  align="R")
            pdf.cell(20, 8, txt=f"{artikel.mwstNorm:.2f}".replace('.', ','),  align="R")
            pdf.cell(20, 8, txt=f"{pro100gramm:.2f}".replace('.', ','), ln=1, align="R")
            # zeile += 10

        try:
            pdf.output(bestellung.datum.strftime("%y-%m-%d") + "-Terrabestellung.pdf")
        except:
            pdf.output("Terrabestellung.pdf")


class Test(unittest.TestCase):

    test_artikel = Artikel(12345, "Testäpfel", 10, 1, "KG", 2.34)
    test_artikelliste = [test_artikel, None]
    test_bestellung = Bestellung(test_artikelliste, datetime.date.today())
 #   print(test_bestellung)
